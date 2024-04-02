from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep
from selenium.webdriver.support.wait import WebDriverWait #ilgili driverı bekleten yapı
from selenium.webdriver.support import expected_conditions as ec #beklenen koşullar
from selenium.webdriver.common.action_chains import ActionChains 
import pytest
import openpyxl
from constants.globalConstants import *
import json 

class Test_Odev3:
      def setup_method(self):
        self.driver=webdriver.Chrome()
        self.driver.maximize_window()
        self.driver.get(BASE_URL)

      def teardown_method(self):
        self.driver.quit()

      def test_blank_login(self):
        userNameInput=self.waitForElelemetVisible((By.ID,username_id))
        passwordInput=self.waitForElelemetVisible((By.ID,password_id))
        loginButton=self.waitForElelemetVisible((By.ID,login_button_id))
        actions=ActionChains(self.driver)
        actions.send_keys_to_element(userNameInput,userNameBlank)
        actions.send_keys_to_element(passwordInput,passwordBlank)
        actions.click(loginButton)
        actions.perform() 
        errorMessage=self.waitForElelemetVisible((By.XPATH,errorMessage_blankXpath))
        assert errorMessage.text== errorMessage_blanktext
    
      def test_blank_password_login(self):
        userNameInput=self.waitForElelemetVisible((By.ID,username_id))
        passwordInput=self.waitForElelemetVisible((By.ID,password_id))
        loginButton=self.waitForElelemetVisible((By.ID,login_button_id))
        actions=ActionChains(self.driver)
        actions.send_keys_to_element(userNameInput,userName)
        actions.send_keys_to_element(passwordInput,passwordBlank)
        actions.click(loginButton)
        actions.perform() 
        errorMessage=self.waitForElelemetVisible((By.XPATH,errorMessage_blankPasswordXpath))
        assert errorMessage.text==errorMessage_blankPasswordtext
        
      def test_lockedUser_login(self):
        userNameInput=self.waitForElelemetVisible((By.ID,username_id))
        passwordInput=self.waitForElelemetVisible((By.ID,password_id))
        loginButton=self.waitForElelemetVisible((By.ID,login_button_id))
        actions=ActionChains(self.driver)
        actions.send_keys_to_element(userNameInput,lokedUserName)
        actions.send_keys_to_element(passwordInput,validPassword)
        actions.click(loginButton)
        actions.perform() 
        errorMessage=self.waitForElelemetVisible((By.XPATH,errorMessage_lokedUserdXpath))
        assert errorMessage.text==errorMessage_lokedUsertext

      def test_valid_login(self):
        userNameInput=self.waitForElelemetVisible((By.ID,username_id))
        passwordInput=self.waitForElelemetVisible((By.ID,password_id))
        loginButton=self.waitForElelemetVisible((By.ID,login_button_id))
        actions=ActionChains(self.driver)
        actions.send_keys_to_element(userNameInput,validUserName)
        actions.send_keys_to_element(passwordInput,validPassword)
        actions.click(loginButton)
        actions.perform()
        baslik=self.waitForElelemetVisible((By.XPATH,baslikXpath))
        assert baslik.text==baslikText
        listOfProducts= self.waitForAllElelemetVisible((By.CSS_SELECTOR,productListCss))
        assert len(listOfProducts)==productListLen

      
      def readInvalidDataFromExcel():
        excelFile = openpyxl.load_workbook("data\invalidLogin.xlsx")
        sheet = excelFile["Sheet1"]
        rows = sheet.max_row #kacıncı satıra kadar benim verim var
        data = []
        for i in range(2,rows+1):
            username = sheet.cell(i,1).value # i. satır 1. sutun daki veriler
            password = sheet.cell(i,2).value #i. satır 2. sutun daki veriler
            data.append((username,password))
        return data
    
      @pytest.mark.parametrize("username,password",readInvalidDataFromExcel())  
      def test_invalid_login(self,username,password):
        userNameInput=self.waitForElelemetVisible((By.ID,username_id))
        passwordInput=self.waitForElelemetVisible((By.ID,password_id))
        loginButton=self.waitForElelemetVisible((By.ID,login_button_id))
        actions=ActionChains(self.driver)
        actions.send_keys_to_element(userNameInput,username)
        actions.send_keys_to_element(passwordInput,password)
        actions.click(loginButton)
        actions.perform()
        errorMesssage=self.waitForElelemetVisible((By.XPATH,errorMessage_InvalidXpath))
        assert errorMesssage.text==errorMessage_Invalidtext

      def readInvalidDataFromJSON(json_file_path):
        with open(json_file_path, 'r') as file:
          data = json.load(file)
          invalid_users = data.get('invalid_credentials', [])
          return [(user.get('username'), user.get('password')) for user in invalid_users]
      
      @pytest.mark.parametrize("username,password",readInvalidDataFromJSON("data\invalidcredentails.json"))  
      def test_invalid_login2(self,username,password):
        userNameInput=self.waitForElelemetVisible((By.ID,username_id))
        passwordInput=self.waitForElelemetVisible((By.ID,password_id))
        loginButton=self.waitForElelemetVisible((By.ID,login_button_id))
        actions=ActionChains(self.driver)
        actions.send_keys_to_element(userNameInput,username)
        actions.send_keys_to_element(passwordInput,password)
        actions.click(loginButton)
        actions.perform()
        errorMesssage=self.waitForElelemetVisible((By.XPATH,errorMessage_InvalidXpath))
        assert errorMesssage.text==errorMessage_Invalidtext


    
      #Sepete başarılı bir şekilde ürün eklenmesinin testi.
      def test_addToCartProduct(self):
        userNameInput=self.waitForElelemetVisible((By.ID,username_id))
        passwordInput=self.waitForElelemetVisible((By.ID,password_id))
        loginButton=self.waitForElelemetVisible((By.ID,login_button_id))
        userNameInput.send_keys(validUserName)
        passwordInput.send_keys(validPassword)
        loginButton.click()
        product = self.waitForElelemetVisible((By.ID,productId))
        actions =ActionChains(self.driver)
        actions.click(product)
        actions.perform()
        shoppingCartBadge = self.waitForElelemetVisible((By.CLASS_NAME,cartBadgeClasName))
        assert shoppingCartBadge.text==cartBadgeText
        
      #Sepete eklenen ürünün başarıyla silinmesi testi
      def test_removeProduct(self):
        userNameInput=self.waitForElelemetVisible((By.ID,username_id))
        passwordInput=self.waitForElelemetVisible((By.ID,password_id))
        loginButton=self.waitForElelemetVisible((By.ID,login_button_id))
        userNameInput.send_keys(validUserName)
        passwordInput.send_keys(validPassword)
        loginButton.click()
        product1 = self.waitForElelemetVisible((By.ID,productId))
        actions =ActionChains(self.driver)
        actions.click(product1)
        actions.perform()
        shoppingCartBadge = self.waitForElelemetVisible((By.CLASS_NAME,cartBadgeClasName))
        shoppingCartBadge.click()
        productRemove = self.waitForElelemetVisible((By.ID,removeProductId))
        productRemove.click()
        shoppingCartBadge = self.waitForElelemetInvisible((By.CLASS_NAME,cartBadgeClasName))
        assert shoppingCartBadge

      #Satın alma işlemini başarılı bir şekilde tamamlayıp, teşekkür mesajının görüntülenmesi testi.
      def test_checkoutProduct(self):
        userNameInput=self.waitForElelemetVisible((By.ID,username_id))
        passwordInput=self.waitForElelemetVisible((By.ID,password_id))
        loginButton=self.waitForElelemetVisible((By.ID,login_button_id))
        userNameInput.send_keys(validUserName)
        passwordInput.send_keys(validPassword)
        loginButton.click()
        product1 = self.waitForElelemetVisible((By.ID,productId))
        actions =ActionChains(self.driver)
        actions.click(product1)
        actions.perform()
        shoppingCartBadge = self.waitForElelemetVisible((By.CLASS_NAME,cartBadgeClasName))
        shoppingCartBadge.click()
        checkoutButton = self.waitForElelemetVisible((By.ID,checkoutID))
        checkoutButton.click()
        firstNameBox = self.waitForElelemetVisible((By.ID,firstNameID))
        lasttNameBox = self.waitForElelemetVisible((By.ID,lastNameID))
        postaCodeBox = self.waitForElelemetVisible((By.ID,postaCodeID))
        firstNameBox.send_keys("Emine")
        lasttNameBox.send_keys("Coskun")
        postaCodeBox.send_keys("34854")
        continueButton = self.waitForElelemetVisible((By.ID,continueButtonID))
        continueButton.click()
        finishButton = self.waitForElelemetVisible((By.ID,finishButtonID))
        finishButton.click()
        messageSuccesfull = self.waitForElelemetVisible((By.XPATH,messageSuccesfullXpath))
        assert messageSuccesfull.text == messageSuccesfullText

      def waitForElelemetVisible(self,locator,timeout=5):
        return WebDriverWait(self.driver,timeout).until(ec.visibility_of_element_located(locator))
      def waitForAllElelemetVisible(self,locators,timeout=5):
        return WebDriverWait(self.driver,timeout).until(ec.visibility_of_all_elements_located(locators))
      def waitForElelemetInvisible(self,locator,timeout=5):
        return WebDriverWait(self.driver,timeout).until(ec.invisibility_of_element_located(locator))
